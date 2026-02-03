import os
import json
import re
import zipfile
import xml.etree.ElementTree as ET
import tomllib
import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import threading
import requests
from datetime import datetime
import time

# ---------------- Dependency Status Checker ---------------- #

class DependencyChecker:
    """Check dependency status using package registries"""

    @staticmethod
    def check_npm_package(package_name, current_version):
        """Check NPM package status"""
        try:
            resp = requests.get(f"https://registry.npmjs.org/{package_name}", timeout=5)
            if resp.status_code == 200:
                data = resp.json()
                dist_tags = data.get("dist-tags", {}) or {}
                latest_version = dist_tags.get("latest", "unknown")

                # Try to detect deprecation from the latest version metadata
                versions = data.get("versions", {}) or {}
                latest_meta = versions.get(latest_version, {})
                is_deprecated = bool(latest_meta.get("deprecated")) or bool(data.get("deprecated"))

                if is_deprecated:
                    return "deprecated", latest_version
                if current_version == latest_version:
                    return "latest", latest_version
                return "old", latest_version
        except Exception:
            pass
        return "unknown", "unknown"

    @staticmethod
    def check_pypi_package(package_name, current_version):
        """Check PyPI package status"""
        try:
            resp = requests.get(f"https://pypi.org/pypi/{package_name}/json", timeout=5)
            if resp.status_code == 200:
                data = resp.json()
                latest_version = data.get("info", {}).get("version", "unknown") or "unknown"
                if latest_version == "unknown":
                    return "unknown", "unknown"
                if current_version == latest_version:
                    return "latest", latest_version
                return "old", latest_version
        except Exception:
            pass
        return "unknown", "unknown"

    @staticmethod
    def check_maven_package(package_id, current_version):
        """
        Check Maven package status using Maven Central search API.
        package_id must be 'groupId:artifactId'.
        """
        try:
            if ":" not in package_id:
                return "unknown", "unknown"
            group_id, artifact_id = package_id.split(":", 1)
            resp = requests.get(
                "https://search.maven.org/solrsearch/select",
                params={"q": f"g:{group_id} AND a:{artifact_id}", "rows": 1, "wt": "json"},
                timeout=5
            )
            if resp.status_code == 200:
                data = resp.json()
                docs = (data.get("response") or {}).get("docs") or []
                if docs:
                    latest_version = docs[0].get("latestVersion", "unknown") or "unknown"
                    if latest_version == "unknown":
                        return "unknown", "unknown"
                    if current_version == latest_version:
                        return "latest", latest_version
                    return "old", latest_version
        except Exception:
            pass
        return "unknown", "unknown"

# ---------------- Dependency Parsers ---------------- #

def parse_package_json(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
        deps = {}
        all_deps = {**data.get("dependencies", {}), **data.get("devDependencies", {})}

        for name, version in all_deps.items():
            # Remove ^ ~ >= <= < > * at start
            clean_version = re.sub(r'^[\^~>=<\*]+', '', str(version))
            if isinstance(version, str) and (version.startswith("http") or version.startswith("git")):
                deps[name] = clean_version or "0.0.0"
            elif clean_version in ("", "*"):
                deps[name] = "0.0.0"
            else:
                deps[name] = clean_version
        return deps

def parse_composer_json(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
        deps = {}
        all_deps = {**data.get("require", {}), **data.get("require-dev", {})}
        for name, version in all_deps.items():
            clean_version = re.sub(r'^[\^~>=<\*]+', '', str(version))
            deps[name] = clean_version if clean_version and clean_version != "*" else "0.0.0"
        return deps

def parse_requirements_txt(filepath):
    deps = {}
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            # Handle common operators
            for op in ["==", ">=", "<=", ">", "<"]:
                if op in line:
                    lib, ver = line.split(op, 1)
                    deps[lib.strip()] = ver.strip() or "0.0.0"
                    break
            else:
                deps[line] = "0.0.0"
    return deps

def parse_pyproject_toml(filepath):
    with open(filepath, "rb") as f:
        data = tomllib.load(f)
        project_deps = data.get("project", {}).get("dependencies", []) or []
        deps = {}
        for dep in project_deps:
            dep = str(dep)
            for op in ["==", ">=", ">", "<=", "<", "~=", "!="]:
                if op in dep:
                    lib, ver = dep.split(op, 1)
                    deps[lib.strip()] = ver.strip() or "0.0.0"
                    break
            else:
                deps[dep] = "0.0.0"
        return deps

def parse_pom_xml(filepath):
    deps = {}
    tree = ET.parse(filepath)
    root = tree.getroot()
    ns = {"m": "http://maven.apache.org/POM/4.0.0"}
    for dep in root.findall(".//m:dependency", ns):
        group_id = dep.find("m:groupId", ns)
        artifact_id = dep.find("m:artifactId", ns)
        version = dep.find("m:version", ns)
        if group_id is not None and artifact_id is not None:
            version_text = (version.text if version is not None else "0.0.0") or "0.0.0"
            # Replace property placeholder like ${...} with 0.0.0
            if version_text.startswith("${") and version_text.endswith("}"):
                version_text = "0.0.0"
            deps[f"{group_id.text}:{artifact_id.text}"] = version_text
    return deps

def parse_build_gradle(filepath):
    """
    Parse simple Gradle notations like:
    implementation 'group:artifact:version'
    api "group:artifact:version"
    """
    deps = {}
    pattern = re.compile(r"(implementation|api|compile)\s+['\"]([^:'\"]+):([^:'\"]+):([^:'\"]+)['\"]")
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            m = pattern.search(line)
            if m:
                group, artifact, version = m.group(2), m.group(3), m.group(4)
                deps[f"{group}:{artifact}"] = version
    return deps

def _read_jar_pom_properties(jar):
    """
    Attempt to read Maven coordinates from META-INF/maven/**/pom.properties.
    Returns (groupId, artifactId, version) or (None, None, None)
    """
    try:
        # Find any pom.properties
        candidates = [p for p in jar.namelist() if p.endswith("pom.properties") and p.startswith("META-INF/maven/")]
        for prop_path in candidates:
            with jar.open(prop_path) as pf:
                content = pf.read().decode("utf-8", errors="ignore")
                gid = re.search(r"^groupId=(.+)$", content, re.MULTILINE)
                aid = re.search(r"^artifactId=(.+)$", content, re.MULTILINE)
                ver = re.search(r"^version=(.+)$", content, re.MULTILINE)
                if gid and aid and ver:
                    return gid.group(1).strip(), aid.group(1).strip(), ver.group(1).strip()
    except Exception:
        pass
    return None, None, None

def _read_jar_manifest_metadata(jar):
    """
    Fallback: read Implementation-Title / Implementation-Version / Implementation-Vendor-Id
    Returns (maybe_groupId, maybe_artifactId, version)
    """
    try:
        manifest_files = [f for f in jar.namelist() if f.upper().endswith("META-INF/MANIFEST.MF")]
        if not manifest_files:
            return None, None, None
        with jar.open(manifest_files[0]) as mf:
            content = mf.read().decode("utf-8", errors="ignore")
            impl_title = None
            impl_version = None
            vendor_id = None
            m = re.search(r"^Implementation-Title:\s*(.+)$", content, re.MULTILINE)
            if m:
                impl_title = m.group(1).strip()
            m = re.search(r"^Implementation-Version:\s*([^\s]+)", content, re.MULTILINE)
            if m:
                impl_version = m.group(1).strip()
            m = re.search(r"^Implementation-Vendor-Id:\s*(.+)$", content, re.MULTILINE)
            if m:
                vendor_id = m.group(1).strip()
            # Attempt to guess coords: use vendor_id as groupId if present, title as artifactId
            return vendor_id, impl_title, impl_version
    except Exception:
        pass
    return None, None, None

def parse_jar_files(project_path):
    """
    Returns a dict mapping library identifier -> version.
    If Maven coordinates are available, the key will be 'groupId:artifactId'.
    Otherwise, falls back to the filename (e.g., mylib-1.2.3.jar).
    """
    deps = {}
    for root, _, files in os.walk(project_path):
        for file in files:
            if not file.endswith(".jar"):
                continue
            jar_path = os.path.join(root, file)
            try:
                with zipfile.ZipFile(jar_path, "r") as jar:
                    gid, aid, ver = _read_jar_pom_properties(jar)
                    if not (gid and aid and ver):
                        # Fallback to manifest
                        mgid, maid, mver = _read_jar_manifest_metadata(jar)
                        gid = gid or mgid
                        aid = aid or maid
                        ver = ver or mver

                    if gid and aid:
                        key = f"{gid}:{aid}"
                        deps[key] = ver or "0.0.0"
                    else:
                        # As a last resort, try to extract version from filename like name-1.2.3.jar
                        base = os.path.splitext(file)[0]
                        m = re.search(r"(.+)-(\d+(?:\.\d+)*[A-Za-z0-9\-\._]*)$", base)
                        if m:
                            name_guess, ver_guess = m.group(1), m.group(2)
                            deps[name_guess] = ver_guess
                        else:
                            deps[file] = "0.0.0"
            except Exception:
                deps[file] = "0.0.0"
    return deps

# ---------------- Main Scanner ---------------- #

def scan_project(project_path, progress_callback=None):
    """
    Scan ONLY:
    - JavaScript (package.json)
    - PHP (composer.json)
    - Python (requirements.txt, pyproject.toml)
    - Java (pom.xml, build.gradle, .jar)
    """
    results = {}
    total_files = sum([len(files) for _, _, files in os.walk(project_path)])
    processed = 0

    for root, _, files in os.walk(project_path):
        for file in files:
            filepath = os.path.join(root, file)
            try:
                if file == "package.json":
                    results.setdefault(("JavaScript", "package.json"), {}).update(parse_package_json(filepath))
                elif file == "composer.json":
                    results.setdefault(("PHP", "composer.json"), {}).update(parse_composer_json(filepath))
                elif file == "requirements.txt":
                    results.setdefault(("Python", "requirements.txt"), {}).update(parse_requirements_txt(filepath))
                elif file == "pyproject.toml":
                    results.setdefault(("Python", "pyproject.toml"), {}).update(parse_pyproject_toml(filepath))
                elif file == "pom.xml":
                    results.setdefault(("Java", "pom.xml"), {}).update(parse_pom_xml(filepath))
                elif file == "build.gradle":
                    results.setdefault(("Java", "build.gradle"), {}).update(parse_build_gradle(filepath))
            except Exception as e:
                print(f"⚠️ Error parsing {filepath}: {e}")

            processed += 1
            if progress_callback:
                progress_callback(processed, total_files)

    # Scan .jar files separately - use consistent 2-tuple format
    jar_deps = parse_jar_files(project_path)
    if jar_deps:
        results.setdefault(("Java", "JAR files"), {}).update(jar_deps)

    return results

def check_dependency_status(lang_platform_file, lib, version):
    """Check the status of a dependency"""
    checker = DependencyChecker()
    lang = lang_platform_file[0] if isinstance(lang_platform_file, tuple) else lang_platform_file

    if lang == "JavaScript":
        status, latest = checker.check_npm_package(lib, version)
        if version == "0.0.0" and latest != "unknown":
            return status, latest
        return status, version

    if lang == "Python":
        status, latest = checker.check_pypi_package(lib, version)
        if version == "0.0.0" and latest != "unknown":
            return status, latest
        return status, version

    if lang == "Java":
        status, latest = checker.check_maven_package(lib, version)
        if version == "0.0.0" and latest != "unknown":
            return status, latest
        return status, version

    if lang == "PHP":
        # PHP packages can't be easily checked via public APIs, so return unknown
        return "unknown", version

    return "unknown", version

def save_to_excel(dependencies, output_path="project_dependencies.xlsx", check_status=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dependencies"

    headers = ["Language", "File Type", "Library", "Current Version", "Status", "Last Checked"]
    ws.append(headers)

    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_fills = {
        "latest": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "old": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
        "deprecated": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        "unknown": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    }

    row_num = 2
    for lang_platform_file, libs in dependencies.items():
        for lib, version in libs.items():
            status = "unknown"
            final_version = version

            if check_status:
                status, maybe_ver = check_dependency_status(lang_platform_file, lib, version)
                # adopt discovered version if it differs and is meaningful
                if maybe_ver not in ("unknown", None) and maybe_ver != version:
                    final_version = maybe_ver

            # Handle different tuple formats safely
            if isinstance(lang_platform_file, tuple):
                if len(lang_platform_file) == 3:
                    # Format: (language, platform, file_type)
                    language, platform, file_type = lang_platform_file
                elif len(lang_platform_file) == 2:
                    # Format: (language, file_type) - most common case
                    language, file_type = lang_platform_file
                    platform = ""
                else:
                    # Unexpected format, handle gracefully
                    language = str(lang_platform_file[0]) if lang_platform_file else "Unknown"
                    file_type = str(lang_platform_file[1]) if len(lang_platform_file) > 1 else "Unknown"
                    platform = ""
            else:
                # Fallback for non-tuple format (shouldn't happen in current code)
                language = str(lang_platform_file)
                platform = ""
                file_type = ""

            ws.append([
                language,
                file_type,
                lib,
                final_version,
                status,
                datetime.now().strftime("%Y-%m-%d %H:%M")
            ])

            if status in status_fills:
                ws[f"E{row_num}"].fill = status_fills[status]  # Status is now column E
            row_num += 1

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(output_path)

# ---------------- Loading Screen ---------------- #

class LoadingScreen(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Scanning...")
        self.geometry("400x250")
        self.resizable(False, False)

        self.transient(parent)
        self.grab_set()

        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (400 // 2)
        y = (self.winfo_screenheight() // 2) - (250 // 2)
        self.geometry(f"+{x}+{y}")

        try:
            self.title_font = ("Lora", 18, "bold")
            self.text_font = ("Poppins", 12)
        except Exception:
            self.title_font = ("Arial", 18, "bold")
            self.text_font = ("Arial", 12)

        self.loading_label = ctk.CTkLabel(self, text="🔍 Scanning Projects", font=self.title_font)
        self.loading_label.pack(pady=30)

        self.progress = ctk.CTkProgressBar(self, width=300)
        self.progress.pack(pady=20)
        self.progress.set(0)

        self.status_label = ctk.CTkLabel(self, text="Initializing scanner...", font=self.text_font)
        self.status_label.pack(pady=10)

        self.files_label = ctk.CTkLabel(self, text="Files processed: 0", font=self.text_font)
        self.files_label.pack(pady=5)

    def update_progress(self, value, status_text="", files_text=""):
        self.progress.set(value)
        if status_text:
            self.status_label.configure(text=status_text)
        if files_text:
            self.files_label.configure(text=files_text)
        self.update()

# ---------------- Main GUI ---------------- #

class DependencyScannerGUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("📦 Multi-Language Library Scanner")
        self.geometry("1000x700")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (1000 // 2)
        y = (self.winfo_screenheight() // 2) - (700 // 2)
        self.geometry(f"1000x700+{x}+{y}")

        try:
            self.title_font = ("Lora", 28, "bold")
            self.subtitle_font = ("Lora", 16, "bold")
            self.button_font = ("Poppins", 13, "bold")
            self.text_font = ("Poppins", 12)
        except Exception:
            self.title_font = ("Arial", 28, "bold")
            self.subtitle_font = ("Arial", 16, "bold")
            self.button_font = ("Arial", 13, "bold")
            self.text_font = ("Arial", 12)

        self.projects = []
        self.setup_ui()

    def setup_ui(self):
        main_frame = ctk.CTkFrame(self, corner_radius=0)
        main_frame.pack(fill="both", expand=True, padx=0, pady=0)

        header_frame = ctk.CTkFrame(main_frame, height=100, fg_color="#2E4057")
        header_frame.pack(fill="x", padx=0, pady=0)
        header_frame.pack_propagate(False)

        title_container = ctk.CTkFrame(header_frame, fg_color="transparent")
        title_container.pack(pady=25)

        self.title_label = ctk.CTkLabel(
            title_container, text="📦 Library Scanner", font=self.title_font, text_color="#FFFFFF"
        )
        self.title_label.pack(side="left", padx=10)

        subtitle = ctk.CTkLabel(
            header_frame, text="JavaScript • Python • Java • PHP", font=self.text_font, text_color="#B8C5D6"
        )
        subtitle.pack(pady=(0, 10))

        control_frame = ctk.CTkFrame(main_frame, fg_color="#1A1F2E", height=120)
        control_frame.pack(fill="x", padx=20, pady=(20, 10))
        control_frame.pack_propagate(False)

        self.project_counter = ctk.CTkLabel(
            control_frame, text="Projects Added: 0", font=self.subtitle_font, text_color="#7FB3D5"
        )
        self.project_counter.pack(pady=(15, 10))

        button_container = ctk.CTkFrame(control_frame, fg_color="transparent")
        button_container.pack(pady=10)

        button_style = {"font": self.button_font, "height": 40, "width": 150, "corner_radius": 8, "hover_color": "#3A506B"}

        self.add_button = ctk.CTkButton(
            button_container, text="➕ Add Project", command=self.add_project, fg_color="#4A90E2", **button_style
        )
        self.add_button.grid(row=0, column=0, padx=10)

        self.scan_button = ctk.CTkButton(
            button_container, text="🔍 Scan All", command=self.scan_projects, fg_color="#5CB85C", **button_style
        )
        self.scan_button.grid(row=0, column=1, padx=10)

        self.export_button = ctk.CTkButton(
            button_container, text="💾 Export Excel", command=self.export_excel, fg_color="#F39C12", **button_style
        )
        self.export_button.grid(row=0, column=2, padx=10)

        self.clear_button = ctk.CTkButton(
            button_container, text="🗑️ Clear All", command=self.clear_all, fg_color="#E74C3C", **button_style
        )
        self.clear_button.grid(row=0, column=3, padx=10)

        options_frame = ctk.CTkFrame(main_frame, fg_color="#1A1F2E")
        options_frame.pack(fill="x", padx=20, pady=10)

        self.check_status_var = ctk.BooleanVar(value=False)
        self.check_status_cb = ctk.CTkCheckBox(
            options_frame,
            text="Check dependency status (requires internet)",
            variable=self.check_status_var,
            font=self.text_font,
            checkbox_height=20,
            checkbox_width=20
        )
        self.check_status_cb.pack(pady=10)

        results_container = ctk.CTkFrame(main_frame, fg_color="#1A1F2E")
        results_container.pack(fill="both", expand=True, padx=20, pady=(10, 20))

        results_header = ctk.CTkLabel(
            results_container, text="📊 Scan Results", font=self.subtitle_font, text_color="#7FB3D5"
        )
        results_header.pack(pady=(10, 5))

        self.results_box = ctk.CTkTextbox(
            results_container,
            width=900,
            height=350,
            font=self.text_font,
            fg_color="#0F1419",
            text_color="#E0E0E0",
            scrollbar_button_color="#2E4057",
            scrollbar_button_hover_color="#3A506B"
        )
        self.results_box.pack(pady=10, padx=10, fill="both", expand=True)

        footer = ctk.CTkLabel(
            main_frame,
            text="© 2025 Dependency Scanner | Analyze • Track • Update",
            font=("Poppins", 10),
            text_color="#7B8794"
        )
        footer.pack(pady=10)

    def add_project(self):
        folder = filedialog.askdirectory(title="Select Project Folder")
        if folder:
            self.projects.append(folder)
            self.project_counter.configure(text=f"Projects Added: {len(self.projects)}")
            self.results_box.insert("end", f"✅ Added: {os.path.basename(folder)}\n")
            self.results_box.insert("end", f"   Path: {folder}\n\n")

    def scan_projects(self):
        if not self.projects:
            messagebox.showwarning("No Projects", "Please add at least one project to scan.")
            return

        self.results_box.delete("1.0", "end")
        self.all_dependencies = {}

        loading = LoadingScreen(self)
        loading.update()

        def scan_thread():
            total_projects = len(self.projects)
            for i, project in enumerate(self.projects):
                project_name = os.path.basename(project)

                loading.update_progress(
                    (i / max(total_projects, 1)) * 0.8,
                    f"Scanning: {project_name}",
                    f"Project {i+1} of {total_projects}"
                )

                deps = scan_project(project)

                self.after(0, self.update_results, project_name, deps)

                for lang_platform_file, libs in deps.items():
                    self.all_dependencies.setdefault(lang_platform_file, {}).update(libs)

                time.sleep(0.05)

            if self.check_status_var.get():
                loading.update_progress(0.9, "Checking dependency status...", "Please wait...")

            loading.update_progress(1.0, "Scan Complete!", f"Found {sum(len(libs) for libs in self.all_dependencies.values())} dependencies")
            time.sleep(0.6)
            loading.destroy()

            self.after(0, self.show_completion)

        thread = threading.Thread(target=scan_thread, daemon=True)
        thread.start()

    def update_results(self, project_name, deps):
        self.results_box.insert("end", f"📁 {project_name}\n")
        total_deps = sum(len(libs) for libs in deps.values())
        self.results_box.insert("end", f"   Found {total_deps} dependencies\n")
        
        for lang_platform_file, libs in deps.items():
            if libs:
                
                if isinstance(lang_platform_file, tuple):
                    if len(lang_platform_file) == 3:
                        
                        language, platform, file_type = lang_platform_file
                        display_text = f"{language} - {platform} ({file_type})"
                    elif len(lang_platform_file) == 2:
                        
                        language, file_type = lang_platform_file
                        display_text = f"{language} ({file_type})"
                    else:
                        
                        display_text = str(lang_platform_file)
                else:
                    
                    display_text = str(lang_platform_file)
                
                self.results_box.insert("end", f"   • {display_text}: {len(libs)} packages\n")
        
        self.results_box.insert("end", "\n")
        self.results_box.see("end")

    def show_completion(self):
        total_deps = sum(len(libs) for libs in self.all_dependencies.values())
        self.results_box.insert("end", "─" * 50 + "\n")
        self.results_box.insert("end", f"🎉 SCAN COMPLETE!\n")
        self.results_box.insert("end", f"📊 Total Dependencies: {total_deps}\n")
        self.results_box.insert("end", f"🔧 Languages Found: {len(self.all_dependencies)}\n")
        self.results_box.see("end")

    def export_excel(self):
        if not hasattr(self, "all_dependencies") or not self.all_dependencies:
            messagebox.showerror("Error", "No scanned results to export.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"dependencies_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if file_path:
            try:
                save_to_excel(
                    self.all_dependencies,
                    file_path,
                    check_status=self.check_status_var.get()
                )
                messagebox.showinfo("Export Complete", f"✅ Results saved to:\n{os.path.basename(file_path)}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export: {str(e)}")

    def clear_all(self):
        if self.projects:
            if messagebox.askyesno("Clear All", "Are you sure you want to clear all projects and results?"):
                self.projects = []
                self.project_counter.configure(text="Projects Added: 0")
                self.results_box.delete("1.0", "end")
                if hasattr(self, "all_dependencies"):
                    self.all_dependencies = {}
                self.results_box.insert("end", "🗑️ All data cleared.\n")

# ---------------- Run App ---------------- #

if __name__ == "__main__":
    app = DependencyScannerGUI()
    app.mainloop()
